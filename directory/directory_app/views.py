from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, UpdateView
from django.contrib.auth.mixins import PermissionRequiredMixin

from openpyxl import Workbook, load_workbook

from .models import *


def main(request):
    return render(request, 'Главная.html')


class Organizations(DetailView):
    model = Organization
    template_name = 'Организация №1.html'
    context_object_name = 'org'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['workers'] = User.objects.filter(organization_id=self.kwargs['pk'])
        return context


class UpdateWorker(PermissionRequiredMixin, UpdateView):
    model = User
    template_name = 'update.html'
    context_object_name = 'worker'
    fields = ['name', 'post', 'address', 'communications', 'organization', 'email', 'password', 'role', 'is_superuser']

    def has_permission(self):
        if User.objects.get(id=self.kwargs['pk']).organization.pk == self.request.user.organization.pk and \
                self.request.user.role != 'Пользователь' or self.request.user.is_superuser:
            # Можно редактировать если пользователь является администратором
            # организации или главным администратором
            return True
        return False

    def get_success_url(self):
        return reverse_lazy('main')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['main_admin'] = self.request.user.is_superuser
        context['admin'] = User.objects.get(id=self.kwargs['pk']).organization.pk == self.request.user.organization.pk
        return context


class Department(ListView):
    model = User
    template_name = 'Отдел №1.html'
    context_object_name = 'dep'


def search(request):
    return render(request, 'Поиск.htm')


def search_official(request):
    return render(request, 'Управление №1.html')


def management(request):
    return render(request, 'Управление № 1.html')


def export_to_excel(request):
    queryset = User.objects.all()

    wb = Workbook()
    ws = wb.active

    headers = [field.name for field in queryset.model._meta.fields]
    ws.append(headers)

    for obj in queryset:
        row = [str(getattr(obj, field)) for field in headers]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    wb.save(response)

    return response


def import_from_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('excel_file')
        if file:
            wb = load_workbook(filename=file)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)

            headers = next(rows)
            model_fields = [field.name for field in User._meta.fields]

            for row in rows:
                data = dict(zip(headers, row))

                obj = User()
                for field in model_fields:
                    if field in data:
                        if field == 'organization':
                            organization_name = data[field]
                            if organization_name:
                                organization, created = Organization.objects.get_or_create(name=organization_name)
                                setattr(obj, field, organization)
                        else:
                            setattr(obj, field, data[field])
                obj.save()

            return render(request, 'import_success.html')

    return render(request, 'import.html')
